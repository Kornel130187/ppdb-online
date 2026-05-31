from flask import Flask, request
import random
import csv
import os
from openpyxl import Workbook, load_workbook
import requests

app = Flask(__name__)

pendaftar = []

@app.route('/')
def home():
    return f'''
    <html>
    <head>
    <style>
    body{{
        background-image: url('/static/sekolah.png.png');
        background-size: cover;
        background-position: center;
        background-repeat: no-repeat;
        height: 100vh;
    }}

    .container{{
        width: 400px;
        background: rgba(255,255,255,0.85);
        padding: 20px;
        margin: 60px auto;
        border-radius: 15px;
    }}

    input, textarea{{
        width: 100%;
        padding: 10px;
        margin-top: 5px;
        margin-bottom: 10px;
    }}

    button{{
        width: 100%;
        padding: 12px;
        background: darkblue;
        color: white;
        border: none;
        border-radius: 10px;
    }}
    </style>
    </head>

    <body>
    <div class="container">
    <h1>PPDB ONLINE</h1>

    <h3>Formulir Pendaftaran siswa baru</h3>

    <form action="/daftar" method="POST">

    Nama:<br>
    <input type="text" name="nama"><br><br>

    Umur:<br>
    <input type="number" name="umur"><br><br>

    Alamat:<br>
    <textarea name="alamat"></textarea><br><br>

    Nomor HP:<br>
    <input type="text" name="hp"><br><br>

    <button type="submit">Daftar Sekarang</button>

    </form>
    </div>
    </body>
    </html>
    '''

@app.route('/daftar', methods=['POST'])
def daftar():

    nama = request.form['nama']
    umur = request.form['umur']
    alamat = request.form['alamat']
    hp = request.form['hp']

    nomor = random.randint(1000,9999)

    data = {
        'nomor': nomor,
        'nama': nama,
        'umur': umur,
        'alamat': alamat,
        'hp': hp
    }

    pendaftar.append(data)
    excel_file = "data_siswa.xlsx"

    try:
        workbook = load_workbook(excel_file)

    except:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["nomor", "nama", "umur", "alamat", "hp"])
        workbook.save(excel_file)

    workbook = load_workbook(excel_file)

    sheet = workbook.active
    sheet.append([nomor, nama, umur, alamat, hp])
    workbook.save(excel_file)

    file_ada = os.path.isfile('data_siswa.csv')
    with open('data_siswa.csv', mode='a', newline='', encoding='utf-8') as file: 
        writer = csv.writer(file)
        if not file_ada:  
            writer.writerow(['nomor', 'nama', 'umur', 'alamat', 'hp'])
        writer.writerow([nomor, nama, umur, alamat, hp])
        url_sheet = "https://script.google.com/macros/s/AKfycbxEzBr2jrfOoWYmBWhGj7gYDYNqXvjhBFgDmLmz41CnHEwYnv2oyHZ3Mp6UBsTZY1FlLw/exec"

        data_kirim = {
            "nomor": nomor,
            "nama": nama,
            "umur": umur,
            "alamat": alamat,
            "hp": hp
        }

        requests.post(url_sheet, json=data_kirim)

    return f'''
    <h1>Pendaftaran Berhasil</h1>

    <p>Nomor Pendaftaran: <b>{nomor}</b></p>S
    <p>Nama: {nama}</p>
    <p>Umur: {umur}</p>
    <p>Alamat: {alamat}</p>
    <p>HP: {hp}</p>

    <a href="/">Kembali</a>
    '''

if __name__ == '__main__':
    app.run(debug=True)